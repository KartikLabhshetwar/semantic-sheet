from typing import List, Dict, Any, Optional
import openpyxl
from dataclasses import dataclass
from openpyxl.utils import get_column_letter
import logging
import json
import csv
import os

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class CellData:
    """Represents a cell with its metadata."""
    sheet_name: str
    row: int
    column: int
    value: Any
    formula: Optional[str] = None
    address: str = ""
    data_type: str = ""

@dataclass
class SemanticChunk:
    """Represents a semantic chunk of spreadsheet data."""
    content: str
    metadata: Dict[str, Any]
    chunk_type: str  # 'cell', 'row', 'column', 'sheet_summary', 'formula'

class SpreadsheetReader:
    """Reads and parses Excel and CSV files to extract semantic information."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.cell_data: List[CellData] = []
        self.is_csv = file_path.lower().endswith('.csv')
    
    def _sanitize_metadata_for_chroma(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """
        Convert metadata to ChromaDB-compatible format.
        ChromaDB only accepts: str, int, float, bool, or None
        """
        sanitized = {}
        
        for key, value in metadata.items():
            if value is None:
                sanitized[key] = None
            elif isinstance(value, (str, int, float, bool)):
                sanitized[key] = value
            elif isinstance(value, list):
                # Convert list to comma-separated string
                sanitized[key] = ", ".join(str(item) for item in value)
            elif isinstance(value, dict):
                # Convert dict to JSON string
                sanitized[key] = json.dumps(value)
            else:
                # Convert anything else to string
                sanitized[key] = str(value)
        
        return sanitized
    
    def _extract_formula_safely(self, cell) -> Optional[str]:
        """
        Extract formula from a cell using every possible method.
        This method will NEVER throw an AttributeError!
        """
        try:
            # Method 1: Check if cell has data_type = 'f' (formula)
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                for attr_name in ['formula', '_formula', 'f']:
                    try:
                        formula_val = getattr(cell, attr_name, None)
                        if formula_val and str(formula_val).strip() and str(formula_val) != '=':
                            return str(formula_val)
                    except (AttributeError, TypeError):
                        continue
            
            # Method 2: Try direct attribute access with getattr (safe)
            try:
                formula_val = getattr(cell, 'formula', None)
                if formula_val and str(formula_val).strip() and str(formula_val) != '=':
                    return str(formula_val)
            except (AttributeError, TypeError):
                pass
            
            # Method 3: Check if the cell value starts with '=' (indicates formula)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                return cell.value
            
            return None
        
        except Exception as e:
            logger.warning(f"Formula extraction failed for cell {getattr(cell, 'coordinate', 'unknown')}: {e}")
            return None
        
    def load_workbook(self) -> None:
        """Load the Excel workbook or CSV file with formula preservation."""
        try:
            if self.is_csv:
                # CSV files don't need to be "loaded" like workbooks
                logger.info(f"Preparing to read CSV file: {self.file_path}")
            else:
                # Load Excel with data_only=False to preserve formulas and read formulas
                # keep_links=False to avoid issues with external links
                self.workbook = openpyxl.load_workbook(
                    self.file_path, 
                    data_only=False, 
                    keep_vba=False,
                    read_only=False,
                    keep_links=False
                )
                logger.info(f"Successfully loaded workbook with {len(self.workbook.sheetnames)} sheets")
        except Exception as e:
            raise ValueError(f"Failed to load file: {str(e)}")
    
    def extract_cell_data(self) -> List[CellData]:
        """Extract all cell data from the workbook or CSV file."""
        if self.is_csv:
            return self._extract_csv_data()
        else:
            return self._extract_excel_data()
    
    def _extract_csv_data(self) -> List[CellData]:
        """Extract cell data from CSV file."""
        self.cell_data = []
        sheet_name = os.path.basename(self.file_path).replace('.csv', '')
        
        try:
            with open(self.file_path, 'r', encoding='utf-8', newline='') as csvfile:
                # Try to detect delimiter
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.reader(csvfile, delimiter=delimiter)
                
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        if value.strip():  # Only add non-empty cells
                            # Convert column number to letter for consistency
                            col_letter = get_column_letter(col_idx)
                            address = f"{col_letter}{row_idx}"
                            
                            # Try to infer data type
                            data_type = 's'  # default to string
                            try:
                                float(value)
                                data_type = 'n'  # numeric
                            except ValueError:
                                if value.lower() in ['true', 'false']:
                                    data_type = 'b'  # boolean
                            
                            cell_info = CellData(
                                sheet_name=sheet_name,
                                row=row_idx,
                                column=col_idx,
                                value=value,
                                formula=None,  # CSV doesn't have formulas
                                address=address,
                                data_type=data_type
                            )
                            self.cell_data.append(cell_info)
                            
        except Exception as e:
            logger.error(f"Error reading CSV file: {e}")
            raise ValueError(f"Failed to read CSV file: {str(e)}")
        
        return self.cell_data
    
    def _extract_excel_data(self) -> List[CellData]:
        """Extract cell data from Excel workbook."""
        if not self.workbook:
            self.load_workbook()
        
        self.cell_data = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        try:
                            formula = self._extract_formula_safely(cell)

                            if formula:
                                logger.debug(f"Found formula in {cell.coordinate}: {formula}")

                            data_type = ''
                            if hasattr(cell, 'data_type'):
                                data_type = str(cell.data_type)
                            else:
                                if isinstance(cell.value, (int, float)):
                                    data_type = 'n' 
                                elif isinstance(cell.value, str):
                                    data_type = 's' 
                                elif isinstance(cell.value, bool):
                                    data_type = 'b' 
                                else:
                                    data_type = 's'  
                            
                            cell_info = CellData(
                                sheet_name=sheet_name,
                                row=cell.row,
                                column=cell.column,
                                value=cell.value,
                                formula=formula,
                                address=cell.coordinate,
                                data_type=data_type
                            )
                            self.cell_data.append(cell_info)
                            
                        except Exception as e:
                            logger.warning(f"Error processing cell {cell.coordinate} in sheet '{sheet_name}': {e}")
                            cell_info = CellData(
                                sheet_name=sheet_name,
                                row=cell.row,
                                column=cell.column,
                                value=cell.value,
                                formula=None,
                                address=cell.coordinate,
                                data_type='s'
                            )
                            self.cell_data.append(cell_info)
        
        return self.cell_data
    
    def create_semantic_chunks(self) -> List[SemanticChunk]:
        """Create semantic chunks from the extracted cell data."""
        chunks = []
        
        sheets_data = {}
        for cell in self.cell_data:
            if cell.sheet_name not in sheets_data:
                sheets_data[cell.sheet_name] = []
            sheets_data[cell.sheet_name].append(cell)
        
        for sheet_name, cells in sheets_data.items():
            sheet_summary = self._create_enhanced_sheet_summary(sheet_name, cells)
            chunks.append(sheet_summary)
            
            complete_data_chunks = self._create_complete_data_chunks(sheet_name, cells)
            chunks.extend(complete_data_chunks)

            relationship_chunks = self._create_business_relationship_chunks(sheet_name, cells)
            chunks.extend(relationship_chunks)
            
            for cell in cells:
                cell_chunk = self._create_cell_chunk(cell)
                chunks.append(cell_chunk)

                if cell.formula and cell.formula.strip():
                    formula_chunk = self._create_formula_chunk(cell)
                    chunks.append(formula_chunk)
            
            row_chunks = self._create_row_chunks(sheet_name, cells)
            column_chunks = self._create_enhanced_column_chunks(sheet_name, cells)
            chunks.extend(row_chunks)
            chunks.extend(column_chunks)
        
        return chunks
    
    def _create_sheet_summary(self, sheet_name: str, cells: List[CellData]) -> SemanticChunk:
        """Create a summary chunk for a sheet (keeping original method for compatibility)."""
        return self._create_enhanced_sheet_summary(sheet_name, cells)
    
    def _create_cell_chunk(self, cell: CellData) -> SemanticChunk:
        """Create a semantic chunk for a single cell."""
        content = f"In sheet '{cell.sheet_name}', cell {cell.address} contains: {cell.value}"
        
        metadata = self._sanitize_metadata_for_chroma({
            "sheet_name": cell.sheet_name,
            "address": cell.address,
            "row": cell.row,
            "column": cell.column,
            "value": str(cell.value),
            "data_type": cell.data_type
        })
        
        return SemanticChunk(
            content=content,
            metadata=metadata,
            chunk_type="cell"
        )
    
    def _create_formula_chunk(self, cell: CellData) -> SemanticChunk:
        """Create a semantic chunk for a formula."""
        formula_text = cell.formula if cell.formula else "Unknown formula"

        if cell.value is not None:
            content = f"Formula in sheet '{cell.sheet_name}', cell {cell.address}: {formula_text} = {cell.value}"
        else:
            content = f"Formula in sheet '{cell.sheet_name}', cell {cell.address}: {formula_text}"
        
        metadata = self._sanitize_metadata_for_chroma({
            "sheet_name": cell.sheet_name,
            "address": cell.address,
            "formula": formula_text,
            "result": str(cell.value),
            "has_result": cell.value is not None
        })
        
        return SemanticChunk(
            content=content,
            metadata=metadata,
            chunk_type="formula"
        )
    
    def _create_key_value_chunks(self, sheet_name: str, cells: List[CellData]) -> List[SemanticChunk]:
        """Create chunks specifically for key-value pair data like metrics tables."""
        chunks = []
        
        rows = {}
        for cell in cells:
            if cell.row not in rows:
                rows[cell.row] = {}
            rows[cell.row][cell.column] = cell
        
        for row_num, row_cells in rows.items():
            if len(row_cells) >= 2:
                col1_cell = row_cells.get(1)  
                col2_cell = row_cells.get(2)  
                
                if col1_cell and col2_cell and col1_cell.value and col2_cell.value:
                    key = str(col1_cell.value).strip()
                    value = str(col2_cell.value).strip()
                    
                    if key.lower() in ['key metrics', 'metric', 'description', 'name'] or not value:
                        continue
                    
                    content = f"Financial Metric: {key} = {value}"
                    

                    key_lower = key.lower()
                    if any(keyword in key_lower for keyword in ['revenue', 'sales', 'income']):
                        content += " (Revenue/Sales metric)"
                    elif any(keyword in key_lower for keyword in ['profit', 'margin', 'earnings']):
                        content += " (Profitability metric)"
                    elif any(keyword in key_lower for keyword in ['growth', 'rate', '%']):
                        content += " (Growth/Performance metric)"
                    elif any(keyword in key_lower for keyword in ['expense', 'cost', 'cogs', 'cost of goods sold']):
                        content += " (Cost/Expense metric)"

                    try:
                        numeric_value = float(str(value).replace(',', '').replace('$', '').replace('%', ''))
                        if 'margin' in key_lower or '%' in key_lower or 'rate' in key_lower:
                            if numeric_value < 1:
                                content += f" ({numeric_value:.1%} percentage)"
                            else:
                                content += f" ({numeric_value}% percentage)"
                        elif numeric_value >= 1000:
                            content += f" (${numeric_value:,.0f} currency value)"
                    except ValueError:
                        pass
                    
                    metadata = self._sanitize_metadata_for_chroma({
                        "sheet_name": sheet_name,
                        "row": row_num,
                        "metric_name": key,
                        "metric_value": value,
                        "key_column": "A",
                        "value_column": "B"
                    })
                    
                    chunks.append(SemanticChunk(
                        content=content,
                        metadata=metadata,
                        chunk_type="key_value_metric"
                    ))
        if len(rows) > 1:
            header_row = rows.get(1, {})
            time_headers = []
            for col_num in sorted(header_row.keys()):
                if col_num > 1:
                    header_cell = header_row.get(col_num)
                    if header_cell and header_cell.value:
                        header_val = str(header_cell.value).strip()
                        if any(time_word in header_val.lower() for time_word in ['year', 'month', 'quarter', 'period', 'q1', 'q2', 'q3', 'q4']):
                            time_headers.append((col_num, header_val))
            
            if time_headers:
                for row_num, row_cells in rows.items():
                    if row_num == 1:  
                        continue

                    metric_cell = row_cells.get(1)
                    if not metric_cell or not metric_cell.value:
                        continue
                    
                    metric_name = str(metric_cell.value).strip()

                    if metric_name.lower() in ['ratio', 'metric', 'description', 'name']:
                        continue

                    for col_num, time_header in time_headers:
                        value_cell = row_cells.get(col_num)
                        if value_cell and value_cell.value is not None:
                            value = str(value_cell.value).strip()

                            if not value or value == '' or (value == '0' and 'growth' in metric_name.lower()):
                                continue

                            content = f"Financial Ratio: {metric_name} ({time_header}) = {value}"

                            metric_lower = metric_name.lower()
                            if any(keyword in metric_lower for keyword in ['expense', 'cost', 'cogs']):
                                content += " (Cost/Expense ratio)"
                            elif any(keyword in metric_lower for keyword in ['profit', 'margin', 'earnings']):
                                content += " (Profitability ratio)"
                            elif any(keyword in metric_lower for keyword in ['growth', 'rate']):
                                content += " (Growth ratio)"
                            elif any(keyword in metric_lower for keyword in ['revenue', 'sales']):
                                content += " (Revenue ratio)"

                            try:
                                numeric_value = float(value.replace(',', '').replace('$', '').replace('%', ''))
                                if numeric_value < 1 and numeric_value > 0:
                                    content += f" ({numeric_value:.2%} formatted)"
                                elif numeric_value >= 1000:
                                    content += f" (${numeric_value:,.0f} currency)"
                            except ValueError:
                                pass
                            
                            metadata = self._sanitize_metadata_for_chroma({
                                "sheet_name": sheet_name,
                                "row": row_num,
                                "column": col_num,
                                "metric_name": metric_name,
                                "time_period": time_header,
                                "metric_value": value,
                                "table_type": "ratio_table"
                            })
                            
                            chunks.append(SemanticChunk(
                                content=content,
                                metadata=metadata,
                                chunk_type="ratio_metric"
                            ))
        
        return chunks
    
    def _create_row_chunks(self, sheet_name: str, cells: List[CellData]) -> List[SemanticChunk]:
        """Create semantic chunks for rows."""
        chunks = []
        
        rows = {}
        for cell in cells:
            if cell.row not in rows:
                rows[cell.row] = []
            rows[cell.row].append(cell)
        
        for row_num, row_cells in rows.items():
            values = [str(c.value) for c in row_cells if c.value is not None]
            if len(values) > 1: 
                content = f"Row {row_num} in sheet '{sheet_name}' contains: {', '.join(values)}"
                
                metadata = self._sanitize_metadata_for_chroma({
                    "sheet_name": sheet_name,
                    "row": row_num,
                    "cells_count": len(row_cells),
                    "values": values
                })
                
                chunks.append(SemanticChunk(
                    content=content,
                    metadata=metadata,
                    chunk_type="row"
                ))
        
        return chunks
    
    def _create_column_chunks(self, sheet_name: str, cells: List[CellData]) -> List[SemanticChunk]:
        """Create semantic chunks for columns (keeping original method for compatibility)."""
        return self._create_enhanced_column_chunks(sheet_name, cells)
    
    def process_file(self) -> List[SemanticChunk]:
        """Process the entire spreadsheet file and return semantic chunks."""
        self.extract_cell_data()
        return self.create_semantic_chunks()
    
    def _create_enhanced_sheet_summary(self, sheet_name: str, cells: List[CellData]) -> SemanticChunk:
        """Create an enhanced summary chunk for a sheet with business context."""
        non_empty_cells = [c for c in cells if c.value is not None]
        formulas = [c for c in cells if c.formula and c.formula.strip()]
        
        headers = []
        header_cells = [c for c in cells if c.row == 1 and c.value is not None]
        header_cells.sort(key=lambda x: x.column)
        headers = [str(c.value) for c in header_cells]
        
        content = f"Sheet '{sheet_name}' contains {len(non_empty_cells)} non-empty cells with {len(formulas)} formulas."
        
        if headers:
            content += f" Column headers are: {', '.join(headers)}."

            business_patterns = []
            header_lower = [h.lower() for h in headers]
            
            if any('target' in h and 'actual' in header_lower for h in header_lower):
                business_patterns.append("target vs actual analysis")
            if any('variance' in h or 'difference' in h for h in header_lower):
                business_patterns.append("variance analysis")
            if any('%' in h or 'percent' in h or 'ratio' in h for h in header_lower):
                business_patterns.append("percentage/ratio metrics")
            if any('profit' in h or 'margin' in h for h in header_lower):
                business_patterns.append("profitability analysis")
                
            if business_patterns:
                content += f" This sheet appears to contain: {', '.join(business_patterns)}."
        
        metadata = self._sanitize_metadata_for_chroma({
            "sheet_name": sheet_name,
            "total_cells": len(cells),
            "non_empty_cells": len(non_empty_cells),
            "formulas_count": len(formulas),
            "headers": headers,
            "business_context": ', '.join(business_patterns) if business_patterns else "general data"
        })
        
        return SemanticChunk(
            content=content,
            metadata=metadata,
            chunk_type="enhanced_sheet_summary"
        )
    
    def _create_business_relationship_chunks(self, sheet_name: str, cells: List[CellData]) -> List[SemanticChunk]:
        """Create chunks that capture business relationships between columns."""
        chunks = []

        rows = {}
        for cell in cells:
            if cell.row not in rows:
                rows[cell.row] = {}
            rows[cell.row][cell.column] = cell

        headers = {}
        if 1 in rows:
            for col, cell in rows[1].items():
                if cell.value:
                    headers[col] = str(cell.value).lower()

        target_col = actual_col = variance_col = percent_col = month_col = None
        
        for col, header in headers.items():
            if 'target' in header:
                target_col = col
            elif 'actual' in header:
                actual_col = col
            elif 'variance' in header or 'difference' in header:
                variance_col = col
            elif '%' in header or 'percent' in header:
                percent_col = col
            elif 'month' in header or 'period' in header or 'date' in header:
                month_col = col


        if target_col and actual_col:
            exceeded_targets = []
            month_performance = []
            
            for row_num in range(2, len(rows) + 1):
                if row_num in rows:
                    target_val = rows[row_num].get(target_col, {}).value
                    actual_val = rows[row_num].get(actual_col, {}).value
                    month_val = rows[row_num].get(month_col, {}).value if month_col else f"Row {row_num}"
                    
                    if target_val and actual_val:
                        try:
                            target_num = float(str(target_val).replace(',', '').replace('$', ''))
                            actual_num = float(str(actual_val).replace(',', '').replace('$', ''))
                            
                            performance_status = "exceeded target" if actual_num > target_num else "below target"
                            variance = actual_num - target_num
                            percent_to_target = (actual_num / target_num) * 100 if target_num != 0 else 0
                            
                            month_performance.append(f"{month_val}: Target ${target_num:,.0f}, Actual ${actual_num:,.0f}, Variance ${variance:+,.0f} ({percent_to_target:.1f}% to target) - {performance_status}")
                            
                            if actual_num > target_num:
                                exceeded_targets.append(f"{month_val} (Actual: ${actual_num:,.0f} vs Target: ${target_num:,.0f})")
                                
                        except (ValueError, TypeError):
                            continue
            
            # Create comprehensive target vs actual chunk
            if month_performance:
                content = f"TARGET VS ACTUAL PERFORMANCE ANALYSIS in sheet '{sheet_name}': Complete monthly breakdown: {'; '.join(month_performance)}"
                
                chunks.append(SemanticChunk(
                    content=content,
                    metadata=self._sanitize_metadata_for_chroma({
                        "sheet_name": sheet_name,
                        "analysis_type": "target_vs_actual_complete",
                        "target_column": get_column_letter(target_col),
                        "actual_column": get_column_letter(actual_col),
                        "months_analyzed": len(month_performance),
                        "months_exceeded": len(exceeded_targets)
                    }),
                    chunk_type="business_relationship"
                ))

            if exceeded_targets:
                content = f"MONTHS THAT EXCEEDED TARGETS in sheet '{sheet_name}': {', '.join(exceeded_targets)}. These {len(exceeded_targets)} months had actual revenue higher than target revenue."
                
                chunks.append(SemanticChunk(
                    content=content,
                    metadata=self._sanitize_metadata_for_chroma({
                        "sheet_name": sheet_name,
                        "analysis_type": "exceeded_targets",
                        "months_count": len(exceeded_targets)
                    }),
                    chunk_type="exceeded_targets"
                ))


        if target_col and actual_col and variance_col:
            variance_data = []
            for row_num in range(2, len(rows) + 1):
                if row_num in rows:
                    target_val = rows[row_num].get(target_col, {}).value if target_col in rows[row_num] else None
                    actual_val = rows[row_num].get(actual_col, {}).value if actual_col in rows[row_num] else None
                    variance_val = rows[row_num].get(variance_col, {}).value if variance_col in rows[row_num] else None
                    month_val = rows[row_num].get(month_col, {}).value if month_col else f"Row {row_num}"
                    
                    if target_val and actual_val and variance_val:
                        try:
                            variance_num = float(str(variance_val).replace(',', '').replace('$', ''))
                            status = "exceeded target" if variance_num > 0 else "below target"
                            variance_data.append(f"{month_val}: Target {target_val}, Actual {actual_val}, Variance {variance_val} ({status})")
                        except (ValueError, TypeError):
                            continue
            
            if variance_data:
                content = f"VARIANCE ANALYSIS in sheet '{sheet_name}': {headers.get(target_col, 'Target')} vs {headers.get(actual_col, 'Actual')} with calculated {headers.get(variance_col, 'Variance')}. Complete data: {'; '.join(variance_data)}"
                
                chunks.append(SemanticChunk(
                    content=content,
                    metadata=self._sanitize_metadata_for_chroma({
                        "sheet_name": sheet_name,
                        "analysis_type": "variance_analysis",
                        "target_column": get_column_letter(target_col),
                        "actual_column": get_column_letter(actual_col),
                        "variance_column": get_column_letter(variance_col)
                    }),
                    chunk_type="business_relationship"
                ))

        return chunks
    
    def _create_enhanced_column_chunks(self, sheet_name: str, cells: List[CellData]) -> List[SemanticChunk]:
        """Create enhanced semantic chunks for columns with business context."""
        chunks = []

        columns = {}
        for cell in cells:
            col_letter = get_column_letter(cell.column)
            if col_letter not in columns:
                columns[col_letter] = []
            columns[col_letter].append(cell)
        
        for col_letter, col_cells in columns.items():
            col_cells.sort(key=lambda x: x.row) 
            values = [str(c.value) for c in col_cells if c.value is not None]
            
            if len(values) > 1: 
                header = values[0] if col_cells[0].row == 1 else "Unknown"
                data_values = values[1:] if col_cells[0].row == 1 else values

                content = f"Column {col_letter} in sheet '{sheet_name}'"
                if header and header != "Unknown":
                    content += f" ('{header}')"
                content += f" contains {len(data_values)} data values"

                if header:
                    header_lower = header.lower()
                    if 'variance' in header_lower or 'difference' in header_lower:
                        content += " representing variance/difference calculations"
                        try:
                            numeric_values = []
                            for val in data_values[:10]:
                                clean_val = str(val).replace(',', '').replace('$', '').replace('%', '')
                                try:
                                    numeric_values.append(float(clean_val))
                                except ValueError:
                                    continue
                            
                            positive_count = sum(1 for v in numeric_values if v > 0)
                            negative_count = sum(1 for v in numeric_values if v < 0)
                            
                            if positive_count > 0 or negative_count > 0:
                                content += f" with {positive_count} positive and {negative_count} negative values"
                        except:
                            pass
                    
                    elif '%' in header_lower or 'percent' in header_lower or 'ratio' in header_lower:
                        content += " representing percentage/ratio metrics"
                        try:
                            over_100_count = 0
                            for val in data_values[:10]:  # Check first 10 values
                                clean_val = str(val).replace('%', '').replace(',', '')
                                try:
                                    if float(clean_val) > 100 or float(clean_val) > 1.0:
                                        over_100_count += 1
                                except ValueError:
                                    continue
                            
                            if over_100_count > 0:
                                content += f" with {over_100_count} values exceeding 100%/target"
                        except:
                            pass
                    
                    elif 'target' in header_lower:
                        content += " containing target/goal values"
                    elif 'actual' in header_lower:
                        content += " containing actual/achieved values"
                
                content += f". Values include: {', '.join(data_values[:5])}" 
                if len(data_values) > 5:
                    content += f" and {len(data_values) - 5} more"
                
                metadata = self._sanitize_metadata_for_chroma({
                    "sheet_name": sheet_name,
                    "column": col_letter,
                    "header": header,
                    "cells_count": len(col_cells),
                    "data_values_count": len(data_values),
                    "sample_values": data_values[:5]
                })
                
                chunks.append(SemanticChunk(
                    content=content,
                    metadata=metadata,
                    chunk_type="enhanced_column"
                ))
        
        return chunks
    
    def _create_complete_data_chunks(self, sheet_name: str, cells: List[CellData]) -> List[SemanticChunk]:
        """
        Create comprehensive chunks that contain complete column data.
        This solves the retrieval problem by ensuring all data is accessible in single chunks.
        """
        chunks = []
        columns = {}
        for cell in cells:
            col_letter = get_column_letter(cell.column)
            if col_letter not in columns:
                columns[col_letter] = []
            columns[col_letter].append(cell)
        
        for col_letter, col_cells in columns.items():
            col_cells.sort(key=lambda x: x.row)
            
            if len(col_cells) > 1: 
                header = None
                data_values = []

                for cell in col_cells:
                    if cell.row == 1:  # Header row
                        header = str(cell.value) if cell.value else f"Column {col_letter}"
                    else:
                        if cell.value is not None:
                            data_values.append(str(cell.value))
                
                if header and data_values:
                    content = f"COMPLETE COLUMN DATA - Sheet '{sheet_name}', Column {col_letter} ('{header}') contains ALL values: {', '.join(data_values)}"
                    
                    header_lower = header.lower()
                    if any(keyword in header_lower for keyword in ['revenue', 'sales', 'income', 'profit', 'actual', 'target']):
                        content += f". This is a financial data column with {len(data_values)} entries."
                        
                        try:
                            numeric_values = []
                            for i, val in enumerate(data_values):
                                try:
                                    clean_val = float(str(val).replace(',', '').replace('$', ''))
                                    numeric_values.append((clean_val, val, i+2))  # i+2 for row number (1-indexed + header)
                                except ValueError:
                                    continue
                            
                            if numeric_values:
                                numeric_values.sort(key=lambda x: x[0])  # Sort by numeric value
                                lowest = numeric_values[0]
                                highest = numeric_values[-1]
                                
                                content += f" Range: Lowest {lowest[1]} (row {lowest[2]}), Highest {highest[1]} (row {highest[2]})."
                                
                                high_values = [item for item in numeric_values if item[0] > (highest[0] * 0.8)]  # Top 20% values
                                if len(high_values) > 1:
                                    high_desc = [f"{item[1]} (row {item[2]})" for item in high_values]
                                    content += f" High-performing entries: {', '.join(high_desc)}."
                        except Exception:
                            pass
                    
                    metadata = self._sanitize_metadata_for_chroma({
                        "sheet_name": sheet_name,
                        "column": col_letter,
                        "header": header,
                        "data_count": len(data_values),
                        "complete_data": True,
                        "all_values": data_values[:50]
                    })
                    
                    chunks.append(SemanticChunk(
                        content=content,
                        metadata=metadata,
                        chunk_type="complete_column_data"
                    ))

        key_value_chunks = self._create_key_value_chunks(sheet_name, cells)
        chunks.extend(key_value_chunks)
        

        rows = {}
        for cell in cells:
            if cell.row not in rows:
                rows[cell.row] = {}
            rows[cell.row][cell.column] = cell

        if 1 in rows:
            headers = {}
            for col, cell in rows[1].items():
                if cell.value:
                    headers[col] = str(cell.value).lower()
            

            time_col = None
            value_cols = []
            
            for col, header in headers.items():
                if any(keyword in header for keyword in ['month', 'date', 'time', 'period']):
                    time_col = col
                elif any(keyword in header for keyword in ['revenue', 'actual', 'target', 'sales', 'value']):
                    value_cols.append((col, header))
            
            if time_col and value_cols:
                for value_col, value_header in value_cols:
                    time_series_data = []
                    
                    for row_num in sorted(rows.keys()):
                        if row_num > 1:
                            time_val = rows[row_num].get(time_col, {}).value if time_col in rows[row_num] else None
                            value_val = rows[row_num].get(value_col, {}).value if value_col in rows[row_num] else None
                            
                            if time_val and value_val:
                                time_series_data.append(f"{time_val}: {value_val}")
                    
                    if time_series_data:
                        content = f"COMPLETE TIME SERIES - Sheet '{sheet_name}', {value_header} over time: {'; '.join(time_series_data)}"
                        
                        metadata = self._sanitize_metadata_for_chroma({
                            "sheet_name": sheet_name,
                            "data_type": "time_series",
                            "value_column": value_header,
                            "complete_data": True,
                            "time_periods": len(time_series_data)
                        })
                        
                        chunks.append(SemanticChunk(
                            content=content,
                            metadata=metadata,
                            chunk_type="complete_time_series"
                        ))
        
        logger.info(f"Created {len(chunks)} complete data chunks for sheet '{sheet_name}'")
        return chunks